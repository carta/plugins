#!/usr/bin/env python3
# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1", "pillow"]
# ///
"""
Generate a Carta-branded XLSX from compensation benchmark data.

Usage:
    uv run export_benchmarks.py --data @/tmp/benchmarks.json --output benchmarks.xlsx \
        --attribution "Data source: Companies with post money valuations between $50M-$100M. Benchmarks released May 2026." \
        [--notional-first]

--data accepts either a JSON string or a @filepath (e.g. @/tmp/data.json).
--notional-first sets equity column order to notional → FD% → shares (peer groups ≥ $500M).
"""

import argparse
import copy
import json
import sys
from pathlib import Path
from openpyxl import load_workbook

ASSETS_DIR = Path(__file__).parent.parent / "assets"
TEMPLATE_PATH = ASSETS_DIR / "carta_benchmarks_template.xlsx"

# The template (carta_benchmarks_template.xlsx) is the brand source of truth —
# it owns title block, subtitle, logo, fonts (SangBleu Versailles for the H1,
# Calibri for body), and the column-header / data-row / attribution styling.
# Brand designers can update the template without touching this script.
#
# Template layout reference:
#   Row 3:  H1 "Compensation Benchmarks" — SangBleu Versailles 22pt, merged A3:H3
#   Row 4:  Subtitle (gray Calibri 11pt, merged A4:L4)
#   Row 6:  Column header row — bold Calibri 12pt, 24 columns (job…equity_notional_p90)
#   Row 7:  Sample data row — used as the styling template for all data rows
#   Row 9:  Attribution row — bold "Data source" label + merged italic gray text
#   Image:  Carta logo anchored in the title area
TEMPLATE_HEADER_ROW = 6
TEMPLATE_DATA_ROW   = 7
TEMPLATE_ATTR_ROW   = 9


# ── Display label / value mappings ───────────────────────────────────────────
# COLUMN_LABELS maps the internal API field name (used as the dict key in each
# row of the input JSON) to the short, human-readable header that appears in
# row 1 of the spreadsheet. Keep this in sync with build_columns().
COLUMN_LABELS = {
    "job":                  "Job",
    "ladder":               "Ladder",
    "level":                "Level",
    "currency":             "Currency",
    "salary_p25":           "Salary P25",
    "salary_p50":           "Salary P50",
    "salary_p75":           "Salary P75",
    "salary_p90":           "Salary P90",
    "tcc_p25":              "TCC P25",
    "tcc_p50":              "TCC P50",
    "tcc_p75":              "TCC P75",
    "tcc_p90":              "TCC P90",
    "equity_fd_pct_p25":    "Equity P25 FD %",
    "equity_fd_pct_p50":    "Equity P50 FD %",
    "equity_fd_pct_p75":    "Equity P75 FD %",
    "equity_fd_pct_p90":    "Equity P90 FD %",
    "equity_shares_p25":    "Equity P25 Shares",
    "equity_shares_p50":    "Equity P50 Shares",
    "equity_shares_p75":    "Equity P75 Shares",
    "equity_shares_p90":    "Equity P90 Shares",
    "equity_notional_p25":  "Equity P25 Notional",
    "equity_notional_p50":  "Equity P50 Notional",
    "equity_notional_p75":  "Equity P75 Notional",
    "equity_notional_p90":  "Equity P90 Notional",
}

# Title Case display mapping for job-area API enums (mirrors the rolematcher
# casing contract — see carta-compensation-rolematcher SKILL.md "Display →
# API enum tables").
JOB_AREA_DISPLAY = {
    "ACCOUNTING":          "Accounting",
    "ADMIN":               "Administrative",
    "CEO":                 "CEO",
    "CORPORATE_AFFAIRS":   "Corporate Affairs",
    "CUSTOMER_SUCCESS":    "Customer Success",
    "DATA":                "Data",
    "DESIGN":              "Design",
    "ENGINEER":            "Engineering",
    "FINANCE":             "Finance",
    "HR":                  "Human Resources",
    "IT":                  "Information Technology",
    "LEGAL":               "Legal",
    "MANUFACTURING":       "Manufacturing",
    "MARKETING":           "Marketing",
    "OPERATIONS":          "Operations",
    "PRODUCT":             "Product",
    "PROJECT_MANAGEMENT":  "Project Management",
    "RESEARCH":            "Research",
    "SALES":               "Sales",
    "STRATEGY":            "Strategy",
    "SUPPORT":             "Support",
    "OTHER":               "Other",
}

# Title Case display mapping for level API enums.
LEVEL_DISPLAY = {
    "ENTRY":     "Entry",
    "MID1":      "Mid 1",
    "MID2":      "Mid 2",
    "SENIOR1":   "Senior 1",
    "SENIOR2":   "Senior 2",
    "STAFF1":    "Staff 1",
    "STAFF2":    "Staff 2",
    "PRINCIPAL": "Principal",
    "VP1":       "VP 1",
    "VP2":       "VP 2",
    "C_LEVEL":   "C-Level",
    "CEO":       "CEO",
    "UNKNOWN":   "Unknown",
}


def display_value(col_name: str, value):
    """Return the user-facing display form for a cell value.

    For the job/level columns, convert API enums (ENGINEER, SENIOR1, …) to
    their Title Case display form (Engineering, Senior 1, …). All other
    columns pass through unchanged.
    """
    if value is None:
        return None
    if col_name == "job":
        return JOB_AREA_DISPLAY.get(value, value)
    if col_name == "level":
        return LEVEL_DISPLAY.get(value, value)
    return value


def build_columns(notional_first: bool) -> list[str]:
    base     = ["job", "ladder", "level", "currency"]
    salary   = ["salary_p25", "salary_p50", "salary_p75", "salary_p90"]
    tcc      = ["tcc_p25", "tcc_p50", "tcc_p75", "tcc_p90"]
    notional = ["equity_notional_p25", "equity_notional_p50", "equity_notional_p75", "equity_notional_p90"]
    fd_pct   = ["equity_fd_pct_p25",   "equity_fd_pct_p50",   "equity_fd_pct_p75",   "equity_fd_pct_p90"]
    shares   = ["equity_shares_p25",   "equity_shares_p50",   "equity_shares_p75",   "equity_shares_p90"]

    equity = (notional + fd_pct + shares) if notional_first else (fd_pct + shares + notional)
    return base + salary + tcc + equity


def _copy_style(src_cell, dst_cell):
    """Copy formatting (font, fill, alignment, border, number_format) from one
    openpyxl cell to another. openpyxl shares style objects across cells via
    reference, so mutating one would corrupt the template's other cells — hence
    the per-attribute copy. `copy.copy` is sufficient (not `deepcopy`) because
    openpyxl's `Serialisable.__copy__` recursively recreates attributes rather
    than sharing references.
    """
    if src_cell.has_style:
        dst_cell.font          = copy.copy(src_cell.font)
        dst_cell.fill          = copy.copy(src_cell.fill)
        dst_cell.border        = copy.copy(src_cell.border)
        dst_cell.alignment     = copy.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection    = copy.copy(src_cell.protection)


def fill_benchmarks_sheet(ws, rows: list[dict], columns: list[str], attribution: str):
    """Populate the Benchmarks sheet of the template with fetched data.

    The template already provides the title block, subtitle, logo, column header
    row (row 6, raw API names), one sample data row (row 7), and the attribution
    row (row 9). We overwrite:
      - Row 6 header labels → our short Title Case labels (Job, Salary P25, …)
      - Row 7+ data → fetched rows (Title-Case display values for job/level)
      - Row N+1 attribution text → the actual peer-group + benchmark-version string

    Styling is preserved by copying from the template's row 6 (header style) and
    row 7 (data style) onto each new cell, so font/fill/alignment match what the
    designer set.
    """
    # ── 1. Overwrite column header labels in row 6 ────────────────────────────
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=TEMPLATE_HEADER_ROW, column=col_idx)
        cell.value = COLUMN_LABELS.get(col_name, col_name)
        # Style is already correct (set by template); leave it alone.

    # ── 2. Capture row 7 (sample data) style per column, then clear it ────────
    sample_style = {}
    for col_idx in range(1, len(columns) + 1):
        sample_style[col_idx] = ws.cell(row=TEMPLATE_DATA_ROW, column=col_idx)
    # We don't actually clear yet — we'll overwrite below.

    # ── 3. Write each fetched data row starting at row 7, copying style ──────
    for row_offset, row in enumerate(rows):
        excel_row = TEMPLATE_DATA_ROW + row_offset
        for col_idx, col_name in enumerate(columns, 1):
            value = display_value(col_name, row.get(col_name))
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            # Mirror the template's row-7 styling onto this cell (font, align, …).
            _copy_style(sample_style[col_idx], cell)

    # ── 4. Update the attribution row ─────────────────────────────────────────
    # The template has the attribution at row 9 (immediately below sample data
    # row 7 + a blank row at 8). With N data rows, the attribution shifts to
    # TEMPLATE_DATA_ROW + N + 1 (one blank row gap).
    new_attr_row = TEMPLATE_DATA_ROW + len(rows) + 1

    if new_attr_row != TEMPLATE_ATTR_ROW:
        # Copy the template's attribution row styling onto the new position,
        # then clear the original row.
        for col_idx in (1, 2):
            src = ws.cell(row=TEMPLATE_ATTR_ROW, column=col_idx)
            dst = ws.cell(row=new_attr_row, column=col_idx)
            _copy_style(src, dst)
            src.value = None  # clear the old position

        # Also relocate the merged range for the attribution text.
        for mr in list(ws.merged_cells.ranges):
            if mr.min_row == TEMPLATE_ATTR_ROW:
                ws.unmerge_cells(str(mr))
                ws.merge_cells(
                    start_row=new_attr_row, start_column=mr.min_col,
                    end_row=new_attr_row,   end_column=mr.max_col,
                )

    # Now write the actual attribution text.
    ws.cell(row=new_attr_row, column=1).value = "Data source"
    ws.cell(row=new_attr_row, column=2).value = attribution


def main():
    parser = argparse.ArgumentParser(description="Export Carta comp benchmarks as branded XLSX")
    parser.add_argument("--data",           required=True, help="JSON rows string or @filepath")
    parser.add_argument("--output",         required=True, help="Output .xlsx path")
    parser.add_argument("--attribution",    required=True, help="Data-source attribution string")
    parser.add_argument("--notional-first", action="store_true",
                        help="Peer group ≥ $500M: put notional equity columns first")
    args = parser.parse_args()

    # Load data
    if args.data.startswith("@"):
        with open(args.data[1:]) as f:
            rows = json.load(f)
    else:
        rows = json.loads(args.data)

    if not isinstance(rows, list):
        print("Error: --data must be a JSON array of row objects", file=sys.stderr)
        sys.exit(1)

    # Empty input is a no-op, not a workbook we should still write. With zero
    # data rows, fill_benchmarks_sheet would shift the attribution row from
    # row 9 up to row 8 (TEMPLATE_DATA_ROW + 0 + 1), then unmerge and clear
    # the template's row 9 — producing a corrupted workbook with the
    # attribution misaligned above where the data would normally appear.
    # Reject early with a clear message so the caller can decide what to do.
    if not rows:
        print("Error: --data is an empty array; no benchmarks to export", file=sys.stderr)
        sys.exit(2)

    columns = build_columns(args.notional_first)

    # Open the brand-managed template — it owns title block, subtitle, logo,
    # fonts (SangBleu Versailles for H1, Calibri for body), column-header and
    # data-row styling, and the attribution row. We only swap in the dynamic
    # bits (header labels, data rows, attribution text).
    if not TEMPLATE_PATH.exists():
        print(f"Error: template not found at {TEMPLATE_PATH}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["Benchmarks"] if "Benchmarks" in wb.sheetnames else wb.active

    fill_benchmarks_sheet(ws, rows, columns, args.attribution)

    wb.save(args.output)
    print(f"Saved: {args.output}")


if __name__ == "__main__":
    main()
